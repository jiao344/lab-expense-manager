Exit code: 0
Wall time: 0.4 seconds
Output:
from __future__ import annotations

import io
import os
import shutil
import sqlite3
from datetime import date, datetime
from pathlib import Path

from flask import (
    Flask, flash, redirect, render_template, request, send_file, url_for
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

APP_NAME = "鏈哄櫒浜虹爺绌舵墍鏃ュ父缁忚垂绠＄悊"
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
BACKUP_DIR = BASE_DIR / "backups"
DB_PATH = DATA_DIR / "expense.db"
ALLOWED_INVOICE = {"宸叉敹绁?, "鏈敹绁?, "寰呰ˉ绁?, "涓嶉渶绁ㄦ嵁"}
ALLOWED_TYPE = {"鏀跺叆", "鏀嚭"}

app = Flask(__name__)
app.config.update(
    SECRET_KEY=os.environ.get("EXPENSE_SECRET_KEY", "local-expense-manager"),
    MAX_CONTENT_LENGTH=8 * 1024 * 1024,
)


def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    BACKUP_DIR.mkdir(exist_ok=True)
    with db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS settings(
                key TEXT PRIMARY KEY, value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS categories(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                budget REAL NOT NULL DEFAULT 0 CHECK(budget >= 0),
                active INTEGER NOT NULL DEFAULT 1
            );
            CREATE TABLE IF NOT EXISTS people(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                role TEXT NOT NULL DEFAULT '鎴愬憳',
                active INTEGER NOT NULL DEFAULT 1
            );
            CREATE TABLE IF NOT EXISTS transactions(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tx_date TEXT NOT NULL,
                tx_type TEXT NOT NULL CHECK(tx_type IN ('鏀跺叆','鏀嚭')),
                amount REAL NOT NULL CHECK(amount > 0),
                purpose TEXT NOT NULL,
                category TEXT NOT NULL,
                user_name TEXT DEFAULT '',
                handler TEXT DEFAULT '',
                pay_method TEXT DEFAULT '',
                invoice_status TEXT DEFAULT '',
                invoice_no TEXT DEFAULT '',
                remark TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS audit_logs(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                action TEXT NOT NULL,
                detail TEXT NOT NULL,
                created_at TEXT NOT NULL
            );
            """
        )
        defaults = {
            "institute_name": "鏈哄櫒浜虹爺绌舵墍",
            "initial_fund": "0",
            "warning_1": "0.8",
            "warning_2": "0.9",
            "warning_3": "1.0",
        }
        conn.executemany(
            "INSERT OR IGNORE INTO settings(key,value) VALUES(?,?)",
            defaults.items(),
        )
        categories = ["娲诲姩", "鑰楁潗", "鍔炲叕", "璁惧", "缁翠慨", "宸梾", "姣旇禌", "鍩硅", "鎺ュ緟", "鍏朵粬"]
        conn.executemany(
            "INSERT OR IGNORE INTO categories(name,budget) VALUES(?,0)",
            [(name,) for name in categories],
        )


def settings_map(conn: sqlite3.Connection | None = None) -> dict[str, str]:
    owns = conn is None
    conn = conn or db()
    result = {row["key"]: row["value"] for row in conn.execute("SELECT key,value FROM settings")}
    if owns:
        conn.close()
    return result


def money(value: str | float | None) -> float:
    try:
        result = round(float(value or 0), 2)
    except (TypeError, ValueError):
        raise ValueError("閲戦鏍煎紡涓嶆纭?)
    if result < 0:
        raise ValueError("閲戦涓嶈兘涓鸿礋鏁?)
    return result


def add_audit(conn: sqlite3.Connection, action: str, detail: str) -> None:
    conn.execute(
        "INSERT INTO audit_logs(action,detail,created_at) VALUES(?,?,?)",
        (action, detail, datetime.now().isoformat(timespec="seconds")),
    )


def summary(conn: sqlite3.Connection) -> dict[str, float | int]:
    s = settings_map(conn)
    initial = money(s.get("initial_fund"))
    row = conn.execute(
        """
        SELECT
          COALESCE(SUM(CASE WHEN tx_type='鏀跺叆' THEN amount END),0) income,
          COALESCE(SUM(CASE WHEN tx_type='鏀嚭' THEN amount END),0) expense,
          COALESCE(SUM(CASE WHEN tx_type='鏀嚭' AND substr(tx_date,1,7)=? THEN amount END),0) month_expense,
          COALESCE(SUM(CASE WHEN tx_type='鏀嚭' AND invoice_status IN ('鏈敹绁?,'寰呰ˉ绁?) THEN 1 ELSE 0 END),0) missing
        FROM transactions
        """,
        (date.today().strftime("%Y-%m"),),
    ).fetchone()
    budget = conn.execute("SELECT COALESCE(SUM(budget),0) value FROM categories").fetchone()["value"]
    balance = initial + row["income"] - row["expense"]
    return {
        "initial": initial,
        "income": row["income"],
        "expense": row["expense"],
        "balance": balance,
        "month_expense": row["month_expense"],
        "missing": row["missing"],
        "budget": budget,
        "budget_rate": (row["expense"] / budget * 100) if budget else 0,
    }


@app.context_processor
def inject_globals():
    return {"app_name": APP_NAME, "today": date.today().isoformat()}


@app.get("/")
def dashboard():
    with db() as conn:
        sm = summary(conn)
        recent = conn.execute(
            "SELECT * FROM transactions ORDER BY tx_date DESC,id DESC LIMIT 8"
        ).fetchall()
        categories = conn.execute(
            """
            SELECT c.name category,c.budget,
              COALESCE(SUM(CASE WHEN t.tx_type='鏀嚭' THEN t.amount END),0) total
            FROM categories c LEFT JOIN transactions t ON t.category=c.name
            WHERE c.active=1 GROUP BY c.id ORDER BY total DESC
            """
        ).fetchall()
        monthly = conn.execute(
            """
            SELECT substr(tx_date,1,7) month,
              SUM(CASE WHEN tx_type='鏀跺叆' THEN amount ELSE 0 END) income,
              SUM(CASE WHEN tx_type='鏀嚭' THEN amount ELSE 0 END) expense
            FROM transactions GROUP BY month ORDER BY month DESC LIMIT 12
            """
        ).fetchall()[::-1]
        people = conn.execute(
            """
            SELECT user_name,COUNT(*) count,SUM(amount) total
            FROM transactions WHERE tx_type='鏀嚭' AND user_name<>''
            GROUP BY user_name ORDER BY total DESC LIMIT 5
            """
        ).fetchall()
        institute = settings_map(conn).get("institute_name", "鏈哄櫒浜虹爺绌舵墍")
    return render_template(
        "dashboard.html", sm=sm, recent=recent, categories=categories,
        monthly=monthly, people=people, institute=institute,
    )


@app.get("/transactions")
def transactions():
    filters = {
        "type": request.args.get("type", "").strip(),
        "category": request.args.get("category", "").strip(),
        "invoice": request.args.get("invoice", "").strip(),
        "start": request.args.get("start", "").strip(),
        "end": request.args.get("end", "").strip(),
        "keyword": request.args.get("keyword", "").strip(),
    }
    sql = "SELECT * FROM transactions WHERE 1=1"
    args: list[str] = []
    mapping = {"type": "tx_type", "category": "category", "invoice": "invoice_status"}
    for key, column in mapping.items():
        if filters[key]:
            sql += f" AND {column}=?"
            args.append(filters[key])
    if filters["start"]:
        sql += " AND tx_date>=?"
        args.append(filters["start"])
    if filters["end"]:
        sql += " AND tx_date<=?"
        args.append(filters["end"])
    if filters["keyword"]:
        sql += " AND (purpose LIKE ? OR remark LIKE ? OR user_name LIKE ? OR handler LIKE ? OR invoice_no LIKE ?)"
        args += [f"%{filters['keyword']}%"] * 5
    sql += " ORDER BY tx_date DESC,id DESC"
    with db() as conn:
        rows = conn.execute(sql, args).fetchall()
        categories = conn.execute("SELECT * FROM categories WHERE active=1 ORDER BY id").fetchall()
        people = conn.execute("SELECT * FROM people WHERE active=1 ORDER BY id").fetchall()
    return render_template(
        "transactions.html", rows=rows, categories=categories,
        people=people, filters=filters, total=sum(r["amount"] for r in rows),
    )


@app.post("/transactions/save")
def save_transaction():
    form = request.form
    try:
        amount = money(form.get("amount"))
        if amount <= 0:
            raise ValueError("閲戦蹇呴』澶т簬 0")
        tx_type = form.get("tx_type", "")
        if tx_type not in ALLOWED_TYPE:
            raise ValueError("鏀舵敮绫诲瀷涓嶆纭?)
        purpose = form.get("purpose", "").strip()
        if not purpose:
            raise ValueError("鐢ㄩ€斾笉鑳戒负绌?)
        invoice = form.get("invoice_status", "")
        if invoice not in ALLOWED_INVOICE:
            invoice = "涓嶉渶绁ㄦ嵁" if tx_type == "鏀跺叆" else "鏈敹绁?
        values = (
            form.get("tx_date") or date.today().isoformat(), tx_type, amount,
            purpose, form.get("category", "鍏朵粬").strip() or "鍏朵粬",
            form.get("user_name", "").strip(), form.get("handler", "").strip(),
            form.get("pay_method", "").strip(), invoice,
            form.get("invoice_no", "").strip(), form.get("remark", "").strip(),
        )
        now = datetime.now().isoformat(timespec="seconds")
        txid = form.get("id", "").strip()
        with db() as conn:
            if txid:
                conn.execute(
                    """
                    UPDATE transactions SET tx_date=?,tx_type=?,amount=?,purpose=?,category=?,
                    user_name=?,handler=?,pay_method=?,invoice_status=?,invoice_no=?,remark=?,updated_at=?
                    WHERE id=?
                    """,
                    values + (now, int(txid)),
                )
                add_audit(conn, "淇敼娴佹按", f"娴佹按 #{txid}锛歿purpose}锛屄amount:.2f}")
                flash("娴佹按璁板綍宸叉洿鏂?, "success")
            else:
                conn.execute(
                    """
                    INSERT INTO transactions(
                      tx_date,tx_type,amount,purpose,category,user_name,handler,pay_method,
                      invoice_status,invoice_no,remark,created_at,updated_at
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    values + (now, now),
                )
                add_audit(conn, "鏂板娴佹按", f"{tx_type}锛歿purpose}锛屄amount:.2f}")
                flash("娴佹按璁板綍宸蹭繚瀛?, "success")
    except (ValueError, sqlite3.Error) as exc:
        flash(str(exc), "error")
    return redirect(url_for("transactions"))


@app.post("/transactions/<int:txid>/delete")
def delete_transaction(txid: int):
    with db() as conn:
        row = conn.execute("SELECT purpose,amount FROM transactions WHERE id=?", (txid,)).fetchone()
        if row:
            conn.execute("DELETE FROM transactions WHERE id=?", (txid,))
            add_audit(conn, "鍒犻櫎娴佹按", f"娴佹按 #{txid}锛歿row['purpose']}锛屄row['amount']:.2f}")
            flash("娴佹按璁板綍宸插垹闄?, "success")
    return redirect(url_for("transactions"))


@app.get("/reports")
def reports():
    with db() as conn:
        sm = summary(conn)
        categories = conn.execute(
            """
            SELECT c.name,c.budget,COUNT(t.id) count,
              COALESCE(SUM(CASE WHEN t.tx_type='鏀嚭' THEN t.amount END),0) total
            FROM categories c LEFT JOIN transactions t ON t.category=c.name
            WHERE c.active=1 GROUP BY c.id ORDER BY total DESC
            """
        ).fetchall()
        people = conn.execute(
            """
            SELECT user_name,COUNT(*) count,SUM(amount) total
            FROM transactions WHERE tx_type='鏀嚭' AND user_name<>''
            GROUP BY user_name ORDER BY total DESC
            """
        ).fetchall()
        months = conn.execute(
            """
            SELECT substr(tx_date,1,7) month,
              SUM(CASE WHEN tx_type='鏀跺叆' THEN amount ELSE 0 END) income,
              SUM(CASE WHEN tx_type='鏀嚭' THEN amount ELSE 0 END) expense
            FROM transactions GROUP BY month ORDER BY month DESC
            """
        ).fetchall()
    return render_template("reports.html", sm=sm, categories=categories, people=people, months=months)


@app.route("/settings", methods=["GET", "POST"])
def settings():
    with db() as conn:
        if request.method == "POST":
            try:
                values = {
                    "institute_name": request.form.get("institute_name", "").strip() or "鏈哄櫒浜虹爺绌舵墍",
                    "initial_fund": str(money(request.form.get("initial_fund"))),
                    "warning_1": str(float(request.form.get("warning_1", "80")) / 100),
                    "warning_2": str(float(request.form.get("warning_2", "90")) / 100),
                    "warning_3": str(float(request.form.get("warning_3", "100")) / 100),
                }
                conn.executemany("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", values.items())
                for row in conn.execute("SELECT id FROM categories"):
                    budget = money(request.form.get(f"budget_{row['id']}", 0))
                    conn.execute("UPDATE categories SET budget=? WHERE id=?", (budget, row["id"]))
                add_audit(conn, "鏇存柊璁剧疆", "鏇存柊鐮旂┒鎵€淇℃伅銆佸垵濮嬬粡璐逛笌棰勭畻閰嶇疆")
                flash("绯荤粺璁剧疆宸蹭繚瀛?, "success")
            except ValueError as exc:
                flash(str(exc), "error")
            return redirect(url_for("settings"))
        categories = conn.execute("SELECT * FROM categories ORDER BY id").fetchall()
        people = conn.execute("SELECT * FROM people ORDER BY id").fetchall()
        logs = conn.execute("SELECT * FROM audit_logs ORDER BY id DESC LIMIT 12").fetchall()
        s = settings_map(conn)
    return render_template("settings.html", s=s, categories=categories, people=people, logs=logs)


@app.post("/settings/category")
def add_category():
    name = request.form.get("name", "").strip()
    if name:
        try:
            with db() as conn:
                conn.execute("INSERT INTO categories(name,budget) VALUES(?,?)", (name, money(request.form.get("budget"))))
                add_audit(conn, "鏂板鍒嗙被", name)
            flash("缁忚垂鍒嗙被宸叉坊鍔?, "success")
        except sqlite3.IntegrityError:
            flash("璇ュ垎绫诲凡瀛樺湪", "error")
    return redirect(url_for("settings"))


@app.post("/settings/person")
def add_person():
    name = request.form.get("name", "").strip()
    if name:
        try:
            with db() as conn:
                conn.execute("INSERT INTO people(name,role) VALUES(?,?)", (name, request.form.get("role", "鎴愬憳").strip()))
                add_audit(conn, "鏂板浜哄憳", name)
            flash("浜哄憳宸叉坊鍔?, "success")
        except sqlite3.IntegrityError:
            flash("璇ヤ汉鍛樺凡瀛樺湪", "error")
    return redirect(url_for("settings"))


@app.get("/export.xlsx")
def export_xlsx():
    with db() as conn:
        rows = conn.execute("SELECT * FROM transactions ORDER BY tx_date,id").fetchall()
        sm = summary(conn)
        institute = settings_map(conn).get("institute_name", "鏈哄櫒浜虹爺绌舵墍")
    wb = Workbook()
    ws = wb.active
    ws.title = "缁忚垂娴佹按"
    ws.merge_cells("A1:M1")
    ws["A1"] = f"{institute}鏃ュ父缁忚垂娴佹按"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="173F5F")
    ws["A1"].alignment = Alignment(horizontal="center")
    headers = ["搴忓彿", "鏃ユ湡", "绫诲瀷", "閲戦锛堝厓锛?, "鐢ㄩ€?, "鍒嗙被", "浣跨敤浜?, "缁忓姙浜?, "鏀粯鏂瑰紡", "绁ㄦ嵁鐘舵€?, "鍙戠エ鍙?, "澶囨敞", "缁撲綑锛堝厓锛?]
    for col, title in enumerate(headers, 1):
        cell = ws.cell(3, col, title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="287271")
        cell.alignment = Alignment(horizontal="center")
    balance = sm["initial"]
    for index, row in enumerate(rows, 1):
        balance += row["amount"] if row["tx_type"] == "鏀跺叆" else -row["amount"]
        values = [index, row["tx_date"], row["tx_type"], row["amount"], row["purpose"], row["category"], row["user_name"], row["handler"], row["pay_method"], row["invoice_status"], row["invoice_no"], row["remark"], balance]
        for col, value in enumerate(values, 1):
            ws.cell(index + 3, col, value)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:M{max(3, len(rows) + 3)}"
    for col in ("D", "M"):
        for cell in ws[col][3:]:
            cell.number_format = '楼#,##0.00'
    widths = [8, 13, 9, 14, 30, 12, 12, 12, 13, 13, 18, 30, 15]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + index)].width = width
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"{APP_NAME}_{date.today()}.xlsx")


@app.post("/import.xlsx")
def import_xlsx():
    upload = request.files.get("file")
    if not upload or not upload.filename.lower().endswith(".xlsx"):
        flash("璇烽€夋嫨 .xlsx 鏂囦欢", "error")
        return redirect(url_for("settings"))
    try:
        workbook = load_workbook(upload, read_only=True, data_only=True)
        sheet = workbook.active
        headers = [str(c.value or "").strip() for c in next(sheet.iter_rows())]
        aliases = {"鏃ユ湡": "tx_date", "绫诲瀷": "tx_type", "閲戦": "amount", "閲戦锛堝厓锛?: "amount", "鐢ㄩ€?: "purpose", "鍒嗙被": "category", "浣跨敤浜?: "user_name", "缁忓姙浜?: "handler", "鏀粯鏂瑰紡": "pay_method", "绁ㄦ嵁鐘舵€?: "invoice_status", "鍙戠エ鍙?: "invoice_no", "澶囨敞": "remark"}
        mapping = {i: aliases[h] for i, h in enumerate(headers) if h in aliases}
        required = {"tx_date", "tx_type", "amount", "purpose"}
        if not required.issubset(mapping.values()):
            raise ValueError("琛ㄥご鑷冲皯闇€瑕侊細鏃ユ湡銆佺被鍨嬨€侀噾棰濄€佺敤閫?)
        count = 0
        now = datetime.now().isoformat(timespec="seconds")
        with db() as conn:
            for row in sheet.iter_rows(values_only=True):
                item = {field: row[index] for index, field in mapping.items()}
                if not any(v not in (None, "") for v in item.values()):
                    continue
                tx_type = str(item.get("tx_type", "")).strip()
                if tx_type not in ALLOWED_TYPE:
                    continue
                tx_date = item.get("tx_date")
                if isinstance(tx_date, datetime):
                    tx_date = tx_date.date().isoformat()
                values = (
                    str(tx_date), tx_type, money(item.get("amount")),
                    str(item.get("purpose", "")).strip(), str(item.get("category", "鍏朵粬") or "鍏朵粬").strip(),
                    str(item.get("user_name", "") or "").strip(), str(item.get("handler", "") or "").strip(),
                    str(item.get("pay_method", "") or "").strip(), str(item.get("invoice_status", "") or "").strip(),
                    str(item.get("invoice_no", "") or "").strip(), str(item.get("remark", "") or "").strip(), now, now,
                )
                conn.execute("INSERT INTO transactions(tx_date,tx_type,amount,purpose,category,user_name,handler,pay_method,invoice_status,invoice_no,remark,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)", values)
                count += 1
            add_audit(conn, "瀵煎叆娴佹按", f"浠?Excel 瀵煎叆 {count} 鏉?)
        flash(f"鎴愬姛瀵煎叆 {count} 鏉℃祦姘?, "success")
    except (ValueError, sqlite3.Error, StopIteration) as exc:
        flash(f"瀵煎叆澶辫触锛歿exc}", "error")
    return redirect(url_for("settings"))


@app.get("/backup")
def backup():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target = BACKUP_DIR / f"expense_backup_{timestamp}.db"
    with db() as source, sqlite3.connect(target) as destination:
        source.backup(destination)
        add_audit(source, "鍒涘缓澶囦唤", target.name)
    return send_file(target, as_attachment=True, download_name=target.name)


@app.post("/restore")
def restore():
    upload = request.files.get("file")
    if not upload or not upload.filename.lower().endswith(".db"):
        flash("璇烽€夋嫨鏈夋晥鐨?.db 澶囦唤鏂囦欢", "error")
        return redirect(url_for("settings"))
    temp = DATA_DIR / "restore_check.db"
    upload.save(temp)
    try:
        with sqlite3.connect(temp) as check:
            tables = {r[0] for r in check.execute("SELECT name FROM sqlite_master WHERE type='table'")}
        if not {"transactions", "settings", "categories", "people"}.issubset(tables):
            raise ValueError("澶囦唤鏂囦欢缁撴瀯涓嶆纭?)
        safety = BACKUP_DIR / f"before_restore_{datetime.now():%Y%m%d_%H%M%S}.db"
        shutil.copy2(DB_PATH, safety)
        os.replace(temp, DB_PATH)
        flash("鏁版嵁宸叉仮澶嶏紝鎭㈠鍓嶇殑鏁版嵁涔熷凡鑷姩澶囦唤", "success")
    except (sqlite3.Error, ValueError) as exc:
        temp.unlink(missing_ok=True)
        flash(f"鎭㈠澶辫触锛歿exc}", "error")
    return redirect(url_for("settings"))


init_db()

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)

